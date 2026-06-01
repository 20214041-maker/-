import streamlit as st
import pandas as pd
import pulp
import calendar
import datetime
import io
import re
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta

st.set_page_config(page_title="성심당 외식사업부 인력 운영 시스템", layout="wide")

st.sidebar.header("⚙️ 시스템 모드")
mode = st.sidebar.radio(
    "운영 범위 선택",
    ["한달 모드", 
     "중장기(순환) 모드",
     "피로도 분석 (최근 스케줄 분석)"]
)

if mode == "한달 모드":
    st.title("🥐 성심당 테라스 스케줄 시스템")
    import streamlit as st
    import pandas as pd
    import pulp
    import calendar
    import datetime
    import io
    import re
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    
    WEEKDAYS_KR  = ['월','화','수','목','금','토','일']
    WEEKDAYS_MAP = {'월':0,'화':1,'수':2,'목':3,'금':4,'토':5,'일':6}
    TIMEVAL      = {'A':1,'B':2,'C':3,'D':4,'E':5}
    TIME_COLORS  = {'A':'FFD9E8FF','B':'FFFFE0B2','C':'FFE8F5E9',
                    'D':'FFFCE4EC','E':'FFFFE9A0','휴':'FFFFFFFF','연':'FFFFFF99'}
    WEEKEND_FILL = PatternFill('solid', fgColor='FFFFC0C0')
    SHORT_FILL   = PatternFill('solid', fgColor='FFFF4444')
    SUM_FILL     = PatternFill('solid', fgColor='FFFFE0B2')
    CENTER = Alignment(horizontal='center', vertical='center')
    THIN   = Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))
    
    # ── 파싱 함수 ────────────────────────────────
    def parse_week_days(text):
        s = set()
        for t in str(text).replace(' ','').split(','):
            if t in WEEKDAYS_MAP: s.add(WEEKDAYS_MAP[t])
        return s
    
    def parse_dates(text):
        s = set()
        if text is None or str(text).strip() in ('', 'nan', 'None'):
            return s
        raw = str(text).strip()
        for token in raw.split(','):
            token = token.strip()
            if not token or token in ('nan','None'): continue
            try:
                val = int(float(token))
                s.add(val)
            except (ValueError, OverflowError):
                pass
        return s
    
    def parse_times(text):
        return [t.strip() for t in str(text).replace(' ','').split(',') if t.strip()]
    
    def parse_num(text, default=5):
        nums = re.findall(r'\d+', str(text))
        return int(nums[0]) if nums else default
    
    # ── 업로드 ──────────────────────────────────
    uploaded_file = st.file_uploader("엑셀 파일을 업로드해주세요.", type=['xlsx'])
    if uploaded_file is None:
        st.stop()
    
    # ── 데이터 로드 ──────────────────────────────
    df_supply   = pd.read_excel(uploaded_file, sheet_name='Supply')
    df_demand   = pd.read_excel(uploaded_file, sheet_name='Demand')
    df_settings = pd.read_excel(uploaded_file, sheet_name='Settings', header=None)
    
    df_supply.columns = [c.strip().replace(' ','') for c in df_supply.columns]
    df_demand.columns = [c.strip().replace(' ','') for c in df_demand.columns]
    df_supply['_ord'] = range(len(df_supply))
    df_supply = df_supply.where(pd.notnull(df_supply), '')
    
    cfg   = dict(zip(df_settings[0].astype(str).str.strip(), df_settings[1]))
    YEAR  = int(cfg.get('연도', 2026))
    MONTH = int(cfg.get('월',   5))
    _, NDAYS = calendar.monthrange(YEAR, MONTH)
    HOLIDAYS = parse_dates(cfg.get('특수공휴일',''))
    DAYS = list(range(1, NDAYS+1))
    
    def is_red(d):
        return datetime.date(YEAR,MONTH,d).weekday() >= 5 or d in HOLIDAYS
    
    RED_DAYS = [d for d in DAYS if is_red(d)]
    
    dc = {}
    for c in df_demand.columns:
        cc = c.replace(' ','')
        if cc=='소속직무':  dc[c]='소속직무'
        elif cc=='타임명':  dc[c]='타임명'
        elif '평일' in cc:  dc[c]='평일'
        elif '주말' in cc:  dc[c]='주말'
    df_demand.rename(columns=dc, inplace=True)
    
    sc = {}
    for c in df_supply.columns:
        cc = c.replace(' ','')
        if cc=='소속직무':                          sc[c]='소속직무'
        elif cc=='가능타임':                        sc[c]='가능타임'
        elif cc=='고정휴무요일':                    sc[c]='고정휴무요일'
        elif '지정휴무일' in cc:                   sc[c]='지정휴무일'
        elif '연차' in cc and cc!='연차':          sc[c]='연차'
        elif cc=='연차':                            sc[c]='연차'
        elif '알바' in cc and '요일' in cc:        sc[c]='알바가능요일'
        elif cc in ('근무일수','주근무일수','주계약일수'): sc[c]='근무일수'
    df_supply.rename(columns=sc, inplace=True)
    
    for col,val in [('고정휴무요일',''),('지정휴무일',''),('연차',''),('알바가능요일',''),('유형','직원')]:
        if col not in df_supply.columns: df_supply[col] = val
    
    # ── 직원 정보 및 타겟 설정 ────────────────────────────────
    emp_info = {}
    target_days = {}
    
    for _, row in df_supply.iterrows():
        name     = str(row['이름'])
        emp_type = str(row.get('유형','직원'))
        avail_t  = parse_times(row.get('가능타임',''))
        off_wday = parse_week_days(row.get('고정휴무요일',''))
        pt_wday  = parse_week_days(row.get('알바가능요일',''))
        off_date = parse_dates(row.get('지정휴무일',''))
        ann_date = parse_dates(row.get('연차',''))
        weekly   = parse_num(row.get('근무일수','주5'), 5)
    
        avail_d = {}
        for d in DAYS:
            wd = datetime.date(YEAR,MONTH,d).weekday()
            ok = True
            if d in off_date or d in ann_date:
                ok = False
            elif emp_type == '알바':
                if pt_wday and wd not in pt_wday: ok = False
            else:
                if wd in off_wday: ok = False
            avail_d[d] = ok
    
        emp_info[name] = {
            'ord':      int(row['_ord']),
            '소속직무': str(row['소속직무']),
            '직급':     str(row['직급']),
            '유형':     emp_type,
            'avail_d':  avail_d,
            'avail_t':  avail_t,
            'weekly':   weekly,
            'ann_date': ann_date,
            'off_date': off_date,
            'is_mgr':   str(row['직급']) in ['계장','대리','팀장'],
            'is_alba':  emp_type == '알바',
        }
    
        target_days[name] = round(weekly * (NDAYS / 7.0))
    
    EMPS  = list(emp_info.keys())
    ROLES = df_supply['소속직무'].unique().tolist()
    TIMES = df_demand['타임명'].unique().tolist()
    
    WEEKS = []
    for week in calendar.monthcalendar(YEAR,MONTH):
        wdays = [d for d in week if d!=0]
        if wdays: WEEKS.append(wdays)
    
    demand = {}
    for _, row in df_demand.iterrows():
        role=str(row['소속직무']); t=str(row['타임명'])
        for d in DAYS:
            demand[(role,d,t)] = int(row['주말']) if is_red(d) else int(row['평일'])
    
    # ════════════════════════════════════════════
    #  사전 인원 충족 검증 (타임별 정밀 검증 깐깐하게!)
    # ════════════════════════════════════════════
    with st.expander("🔍 사전 인원 검증 (스케줄 생성 전 확인)", expanded=True):
        st.markdown("#### 🚨 타임별 정밀 부족 예상 (직무/날짜/타임 기준)")
    
        shortage_preview = []
    
        for d in DAYS:
            dt = datetime.date(YEAR, MONTH, d)
            for role in ROLES:
                r_emps = [i for i in EMPS if emp_info[i]['소속직무'] == role]
    
                for t in TIMES:
                    needed = demand.get((role, d, t), 0)
                    if needed == 0: continue
    
                    avail_cnt = 0
                    for i in r_emps:
                        info = emp_info[i]
                        if info['avail_d'][d] and t in info['avail_t']:
                            avail_cnt += 1
    
                    if avail_cnt < needed:
                        shortage_preview.append({
                            '날짜': f"{d}일({WEEKDAYS_KR[dt.weekday()]})",
                            '빨간날': '🔴' if is_red(d) else '',
                            '직무': role,
                            '타임': t,
                            '필요인원': needed,
                            '가용인원(Max)': avail_cnt,
                            '예상부족': needed - avail_cnt
                        })
    
        if shortage_preview:
            st.error(f"🚨 돌려보나 마나 총 {len(shortage_preview)}건의 타임에서 인원 부족(빵꾸)이 확실시됩니다! 알바생 타임을 조정하거나 대타를 구하세요.")
            df_preview = pd.DataFrame(shortage_preview)
            st.dataframe(df_preview, use_container_width=True, hide_index=True)
            st.caption("※ 가용인원(Max)은 스케줄을 돌리기 전 단순 풀(Pool)입니다. 주5일 제한, 연속근무 금지 등이 걸리면 실제 스케줄 결과의 빵꾸는 이보다 더 많아질 수 있습니다.")
        else:
            st.success("✅ 1차 검증 통과: 일단 모든 타임에 최소 필요 인원 이상의 풀(Pool)이 존재합니다.")
    
    # ════════════════════════════════════════════
    # 스케줄 생성 버튼
    # ════════════════════════════════════════════
    if not st.button("✨ 최적화 스케줄 생성하기"):
        st.stop()
    
    # ════════════════════════════════════════════
    #  1단계: 스케줄 강제 생성 (결원 최소화 최우선)
    # ════════════════════════════════════════════
    with st.spinner("1단계: 스케줄 빈틈없이 채우는 중... (~2분)"):
    
        m1 = pulp.LpProblem("Stage1", pulp.LpMinimize)
    
        X1 = pulp.LpVariable.dicts("x",  (EMPS, DAYS, TIMES), cat='Binary')
        Y1 = pulp.LpVariable.dicts("y",  (EMPS, DAYS),        cat='Binary')
        SH = pulp.LpVariable.dicts("sh", (ROLES, DAYS, TIMES), lowBound=0, cat='Integer')
    
        정규직 = [i for i in EMPS if not emp_info[i]['is_alba']]
        알바생 = [i for i in EMPS if emp_info[i]['is_alba']]
    
        dev_W_정규직 = pulp.LpVariable.dicts("dev_W_regular", 정규직, lowBound=0, cat='Continuous')
        dev_W_알바 = pulp.LpVariable.dicts("dev_W_alba", 알바생, lowBound=0, cat='Continuous')
        max_dev_정규직 = pulp.LpVariable("max_dev_regular", lowBound=0, cat='Continuous')
        max_dev_알바 = pulp.LpVariable("max_dev_alba", lowBound=0, cat='Continuous')
        
        mgr_short = pulp.LpVariable.dicts("mgr_short", DAYS, lowBound=0, cat='Integer')
    
        m1 += (
            100 * pulp.lpSum(SH[r][d][t] for r in ROLES for d in DAYS for t in TIMES)
            + 10 * max_dev_정규직
            + 1 * pulp.lpSum(dev_W_정규직[i] for i in 정규직)
            + 1 * max_dev_알바
            + 0.1 * pulp.lpSum(dev_W_알바[i] for i in 알바생)
            + 0.1 * pulp.lpSum(mgr_short[d] for d in DAYS)
        )
    
        managers = [i for i in EMPS if emp_info[i]['is_mgr']]
    
        for i in EMPS:
            info = emp_info[i]
    
            for d in DAYS:
                if not info['avail_d'][d]:
                    m1 += Y1[i][d] == 0
                    for t in TIMES: m1 += X1[i][d][t] == 0
                else:
                    for t in TIMES:
                        if t not in info['avail_t']:
                            m1 += X1[i][d][t] == 0
    
                m1 += pulp.lpSum(X1[i][d][t] for t in TIMES) == Y1[i][d]
    
            for w_days in WEEKS:
                avail_w = [d for d in w_days if info['avail_d'][d]]
                cap = min(info['weekly'], len(avail_w))
                m1 += pulp.lpSum(Y1[i][d] for d in w_days) <= cap
    
            if 'D' in TIMES and 'A' in TIMES:
                for d in DAYS:
                    if d+1 in DAYS:
                        m1 += X1[i][d]['D'] + X1[i][d+1]['A'] <= 1
    
            for d in range(1, NDAYS - 5):
                m1 += pulp.lpSum(Y1[i][d+k] for k in range(6)) <= 5
    
            tw = pulp.lpSum(Y1[i][d] for d in DAYS) + len(info['ann_date'])
            
            if not info['is_alba']:
                m1 += tw >= target_days[i] - 1
                m1 += tw <= target_days[i] + 1
                
                m1 += dev_W_정규직[i] >= target_days[i] - tw
                m1 += dev_W_정규직[i] >= tw - target_days[i]
                m1 += max_dev_정규직 >= dev_W_정규직[i]
            else:
                m1 += tw >= target_days[i] - 5
                m1 += tw <= target_days[i] + 5
                
                m1 += dev_W_알바[i] >= target_days[i] - tw
                m1 += dev_W_알바[i] >= tw - target_days[i]
                m1 += max_dev_알바 >= dev_W_알바[i]
    
        for role in ROLES:
            r_emps = [i for i in EMPS if emp_info[i]['소속직무']==role]
            for d in DAYS:
                for t in TIMES:
                    needed   = demand.get((role,d,t),0)
                    assigned = pulp.lpSum(X1[i][d][t] for i in r_emps)
                    if needed > 0:
                        m1 += assigned + SH[role][d][t] >= needed
                        m1 += assigned <= needed
                    else:
                        m1 += assigned == 0
    
        if 'D' in TIMES:
            for d in DAYS:
                avail_mgr = [i for i in managers if emp_info[i]['avail_d'][d] and 'D' in emp_info[i]['avail_t']]
                if avail_mgr:
                    m1 += pulp.lpSum(X1[i][d]['D'] for i in avail_mgr) + mgr_short[d] >= 1
    
        m1.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=300, gapRel=0.01))
        stat1 = pulp.LpStatus[m1.status]
    
    assign1 = {}
    for i in EMPS:
        for d in DAYS:
            asgn = '휴'
            for t in TIMES:
                if round(pulp.value(X1[i][d][t]) or 0) == 1:
                    asgn = t; break
            assign1[(i,d)] = asgn
    
    # ════════════════════════════════════════════
    #  2단계: 타임 연속성 개선 (날짜별 스와핑)
    # ════════════════════════════════════════════
    with st.spinner("2단계: 스케줄 연속성 최적화 중..."):
        assign2 = dict(assign1)
    
        def calc_continuity_cost(assign, i):
            info = emp_info[i]
            wds = sorted([d for d in DAYS if assign.get((i,d)) not in (None,'휴')])
            cost = 0
            for idx in range(len(wds)-1):
                d1,d2 = wds[idx],wds[idx+1]
                if d2 != d1+1: continue
                t1,t2 = assign[(i,d1)], assign[(i,d2)]
                if t1 not in TIMEVAL or t2 not in TIMEVAL: continue
                v1,v2 = TIMEVAL[t1], TIMEVAL[t2]
                diff = v2 - v1
                if diff > 0:
                    cost += 20 * diff
                elif diff != 0:
                    cost += 5
            return cost
    
        def calc_total_cost(assign):
            return sum(calc_continuity_cost(assign, i) for i in EMPS)
    
        def is_swap_valid(assign, i, j, d):
            t_i = assign.get((i,d))
            t_j = assign.get((j,d))
            if t_i in (None,'휴','연') or t_j in (None,'휴','연'): return False
            if t_i == t_j: return False
            if t_j not in emp_info[i]['avail_t']: return False
            if t_i not in emp_info[j]['avail_t']: return False
            if d-1 in DAYS and assign.get((i,d-1))=='D' and t_j=='A': return False
            if d+1 in DAYS and t_j=='D' and assign.get((i,d+1))=='A': return False
            if d-1 in DAYS and assign.get((j,d-1))=='D' and t_i=='A': return False
            if d+1 in DAYS and t_i=='D' and assign.get((j,d+1))=='A': return False
            return True
    
        max_iter = 5
        for iteration in range(max_iter):
            improved = False
            current_cost = calc_total_cost(assign2)
    
            for d in DAYS:
                for role in ROLES:
                    r_emps = [i for i in EMPS if emp_info[i]['소속직무']==role
                              and assign2.get((i,d)) not in (None,'휴','연')]
                    for idx_i in range(len(r_emps)):
                        for idx_j in range(idx_i+1, len(r_emps)):
                            i, j = r_emps[idx_i], r_emps[idx_j]
                            if not is_swap_valid(assign2, i, j, d): continue
                            t_i, t_j = assign2[(i,d)], assign2[(j,d)]
                            cost_before = calc_continuity_cost(assign2, i) + calc_continuity_cost(assign2, j)
    
                            assign2[(i,d)], assign2[(j,d)] = t_j, t_i
    
                            demand_ok = True
                            for t in TIMES:
                                needed = demand.get((role, d, t), 0)
                                actual = sum(1 for emp in r_emps if assign2.get((emp, d)) == t)
                                if actual < needed:
                                    demand_ok = False
                                    break
    
                            if demand_ok:
                                cost_after = calc_continuity_cost(assign2, i) + calc_continuity_cost(assign2, j)
                                if cost_after < cost_before:
                                    improved = True
                                else:
                                    assign2[(i,d)], assign2[(j,d)] = t_i, t_j
                            else:
                                assign2[(i,d)], assign2[(j,d)] = t_i, t_j
    
            if not improved:
                break
    
    # ── 결과 수집 ────────────────────────────────
    records = []
    for i in EMPS:
        info = emp_info[i]
        for d in DAYS:
            if d in info['ann_date']:
                asgn = '연'
            else:
                asgn = assign2.get((i,d),'휴')
            records.append({
                'ord':      info['ord'],
                '소속직무': info['소속직무'],
                '직급':     info['직급'],
                '이름':     i,
                '일자':     d,
                '타임':     asgn,
            })
    df_res = pd.DataFrame(records)
    
    shortage_map  = {}
    short_detail  = []
    for role in ROLES:
        r_emps = [i for i in EMPS if emp_info[i]['소속직무']==role]
        for d in DAYS:
            short = False
            for t in TIMES:
                needed = demand.get((role,d,t),0)
                actual = sum(1 for i in r_emps if assign1.get((i,d))==t)
                sv = max(0, needed-actual)
                if sv > 0:
                    short = True
                    short_detail.append({
                        '직무':role,'날짜':d,
                        '요일':WEEKDAYS_KR[datetime.date(YEAR,MONTH,d).weekday()],
                        '타임':t,'필요':needed,'배정':actual,'부족':sv
                    })
            shortage_map[(role,d)] = short
    
    st.success(f"✅ 완료! 1단계: {stat1}")
    
    stat_rows=[]
    for i in EMPS:
        info = emp_info[i]
        ann_cnt = len(info['ann_date'])
        tot = df_res[(df_res['이름']==i)&(~df_res['타임'].isin(['휴','연']))].shape[0]
        we = df_res[(df_res['이름']==i)&(df_res['일자'].isin(RED_DAYS))
                   &(~df_res['타임'].isin(['휴','연']))].shape[0]
        by_t = {t:df_res[(df_res['이름']==i)&(df_res['타임']==t)].shape[0] for t in TIMES}
    
        stat_rows.append({
            '이름': i, '직급': info['직급'], '직무': info['소속직무'], '유형': info['유형'],
            '목표일수': target_days[i], '실제출근': tot, '연차': ann_cnt,
            '총(출근+연차)': tot + ann_cnt,
            '목표대비': (tot + ann_cnt) - target_days[i],
            '주말출근': we, **by_t
        })
    df_stat = pd.DataFrame(stat_rows)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for role in ROLES:
            rdf = df_res[df_res['소속직무']==role].copy()
            pivot = rdf.pivot_table(
                index=['ord','직급','이름'], columns='일자',
                values='타임', aggfunc='first'
            ).reset_index().sort_values('ord').drop(columns=['ord'])
            summary = ['총 근무인원','']
            for d in DAYS:
                summary.append(rdf[(rdf['일자']==d)&(~rdf['타임'].isin(['휴','연']))].shape[0])
            pivot.loc[len(pivot)] = summary
            pivot.to_excel(writer, sheet_name=role, index=False, startrow=3)
            ws = writer.sheets[role]
            ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2+NDAYS)
            tc = ws.cell(1,1,f"{YEAR}년 {MONTH}월 스케줄표 - {role}")
            tc.font=Font(bold=True,size=14); tc.alignment=CENTER
            ws.cell(4,1,'직급').font=Font(bold=True)
            ws.cell(4,2,'이름').font=Font(bold=True)
            for d in DAYS:
                ci=d+2; dt=datetime.date(YEAR,MONTH,d); red=is_red(d)
                for r,val in [(2,d),(3,WEEKDAYS_KR[dt.weekday()])]:
                    cell=ws.cell(r,ci,val)
                    cell.alignment=CENTER; cell.border=THIN
                    if red:
                        cell.font=Font(color='FF0000',bold=True)
                        cell.fill=WEEKEND_FILL
            sum_row=ws.max_row
            for row in ws.iter_rows(min_row=4,max_row=ws.max_row,min_col=1,max_col=2+NDAYS):
                for cell in row:
                    cell.alignment=CENTER; cell.border=THIN
                    val=str(cell.value) if cell.value is not None else ''
                    is_sum=(cell.row==sum_row)
                    if is_sum and cell.column>2:
                        d=cell.column-2
                        if shortage_map.get((role,d),False):
                            cell.fill=SHORT_FILL
                            cell.font=Font(color='FFFFFFFF',bold=True)
                        else:
                            cell.fill=SUM_FILL
                    elif not is_sum and cell.column>2 and val in TIME_COLORS:
                        cell.fill=PatternFill('solid',fgColor=TIME_COLORS[val])
            ws.column_dimensions[get_column_letter(1)].width=7
            ws.column_dimensions[get_column_letter(2)].width=8
            for d in DAYS: ws.column_dimensions[get_column_letter(d+2)].width=3.8
            for r in range(1,ws.max_row+1): ws.row_dimensions[r].height=18
    
        df_stat.to_excel(writer, sheet_name='근무통계', index=False, startrow=1)
        ws_stat = writer.sheets['근무통계']
    
        ws_stat.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_stat.columns))
        title_cell = ws_stat.cell(1, 1, f"{YEAR}년 {MONTH}월 직원별 근무 통계")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = CENTER
    
        HEADER_FILL = PatternFill('solid', fgColor='DDEBF7')
        COLOR_0 = PatternFill('solid', fgColor='E2EFDA')
        COLOR_MINUS = PatternFill('solid', fgColor='BDD7EE')
        COLOR_PLUS = PatternFill('solid', fgColor='FFC7CE')
    
        target_diff_idx = list(df_stat.columns).index('목표대비') + 1
    
        for r_idx, row in enumerate(ws_stat.iter_rows(min_row=2, max_row=ws_stat.max_row, min_col=1, max_col=len(df_stat.columns)), start=2):
            for cell in row:
                cell.alignment = CENTER
                cell.border = THIN
                if r_idx == 2:
                    cell.font = Font(bold=True)
                    cell.fill = HEADER_FILL
                else:
                    if cell.column == target_diff_idx:
                        try:
                            val = int(cell.value)
                            if val == 0: cell.fill = COLOR_0
                            elif val < 0: cell.fill = COLOR_MINUS
                            elif val > 0: cell.fill = COLOR_PLUS
                        except: pass
    
        for col_letter in [get_column_letter(i) for i in range(1, len(df_stat.columns) + 1)]:
            ws_stat.column_dimensions[col_letter].width = 9.5
    
        if short_detail:
            df_short = pd.DataFrame(short_detail)
            df_short.to_excel(writer, sheet_name='인원미달', index=False)
            ws_short = writer.sheets['인원미달']
            for row in ws_short.iter_rows():
                for cell in row:
                    cell.alignment = CENTER
                    cell.border = THIN
                    if cell.row == 1:
                        cell.font = Font(bold=True)
                        cell.fill = HEADER_FILL
            for col_letter in [get_column_letter(i) for i in range(1, len(df_short.columns) + 1)]:
                ws_short.column_dimensions[col_letter].width = 10
    
    st.subheader("📊 직원별 근무 통계")
    st.dataframe(df_stat, use_container_width=True)
    
    if short_detail:
        st.subheader("⚠️ 인원 미달 현황")
        st.dataframe(pd.DataFrame(short_detail), use_container_width=True)
    
        st.subheader("💡 초과근무 요청 가능 직원 추천")
        st.caption("부족한 날 휴무인 직원 중 조건을 만족하는 직원 목록입니다.")
    
        day_to_week = {}
        for w_idx, w_days in enumerate(WEEKS):
            for d in w_days:
                day_to_week[d] = (w_idx, w_days)
    
        for item in short_detail:
            role = item['직무']
            d    = item['날짜']
            t    = item['타임']
            lack = item['부족']
            dt   = datetime.date(YEAR, MONTH, d)
            day_label = f"{d}일({WEEKDAYS_KR[dt.weekday()]})"
    
            with st.expander(f"📅 {day_label} — {role} {t}타임 {lack}명 부족"):
                candidates = []
                for i in EMPS:
                    info = emp_info[i]
                    if info['소속직무'] != role: continue
                    if assign2.get((i, d)) not in ('휴', None): continue
                    if not info['avail_d'][d]: continue
                    if t not in info['avail_t']: continue
                    if t == 'A' and d-1 in DAYS and assign2.get((i, d-1)) == 'D': continue
                    if t == 'D' and d+1 in DAYS and assign2.get((i, d+1)) == 'A': continue
    
                    w_idx, w_days = day_to_week.get(d, (None, []))
                    if w_days:
                        avail_w = [wd for wd in w_days if info['avail_d'][wd]]
                        week_cap = min(info['weekly'], len(avail_w))
                        week_actual = sum(1 for wd in w_days if assign2.get((i, wd)) not in ('휴', None, '연'))
                        has_weekly_room = week_actual < week_cap
                    else:
                        has_weekly_room = True
    
                    month_actual = sum(1 for d2 in DAYS if assign2.get((i, d2)) not in ('휴', None, '연'))
    
                    candidates.append({
                        '이름':       i,
                        '직급':       info['직급'],
                        '유형':       info['유형'],
                        '현재 월근무': month_actual,
                        '목표일수':   target_days[i],
                        '주차 여유':  '✅ 있음' if has_weekly_room else '⚠️ 없음',
                        '권장도':     '★★★' if has_weekly_room and month_actual < target_days[i]
                                      else '★★☆' if has_weekly_room
                                      else '★☆☆'
                    })
    
                if candidates:
                    df_cand = pd.DataFrame(candidates).sort_values('권장도', ascending=False)
                    st.dataframe(df_cand, use_container_width=True, hide_index=True)
                    st.caption("★★★ 주차 여유 + 월 목표 미달  /  ★★☆ 주차 여유 있음  /  ★☆☆ 주차 상한 초과 (초과근무 필요)")
                else:
                    st.info("😔 조건을 만족하는 후보 직원이 없습니다. 외부 인력 투입이 필요합니다.")
    else:
        st.success("🎉 모든 날짜 필요 인원 충족!")
    
    st.download_button("📤 엑셀 다운로드", output.getvalue(),
        f"성심당_테라스_스케줄_{YEAR}년{MONTH}월.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



elif mode == "중장기(순환) 모드":
    st.title("🥐 성심당 외식사업부 인력 운영 최적화 시스템")
    st.header("🌐 사업부 중장기 순환 계획")
    st.info("각 매장×직무의 숙련도 수준(상/중/하)을 설정하고, 적합한 인력을 배치합니다.")
    
    from dateutil.relativedelta import relativedelta
    import numpy as np
    
    def safe_float(value, default=0.0):
        try: 
            return default if pd.isna(value) else float(value)
        except: 
            return default

    def calculate_months_between(start_date, base_date):
        if pd.isna(start_date): return 0
        if isinstance(start_date, str):
            try: start_date = pd.to_datetime(start_date).date()
            except: return 0
        elif isinstance(start_date, datetime.datetime):
            start_date = start_date.date()
        delta = relativedelta(base_date, start_date)
        return delta.years * 12 + delta.months
    
    st.sidebar.subheader("⏱️ 순환 계획 설정")
    
    planning_period_months = st.sidebar.selectbox(
        "전체 계획 기간",
        [6, 12],
        index=1,
        format_func=lambda x: f"{x}개월",
        help="중장기 순환 계획 전체 기간"
    )
    
    first_rotation_month = st.sidebar.selectbox(
        "첫 순환 시점",
        [3, 6],
        index=1,
        format_func=lambda x: f"{x}개월 후",
        help="첫 번째 순환을 몇 개월 후에 시작할지 선택"
    )
    
    min_tenure_months = st.sidebar.selectbox(
        "최소 재직 기간 (순환 조건)",
        [3, 6],
        index=0,
        format_func=lambda x: f"{x}개월",
        help="매장에 최소 이 기간 이상 근무해야 순환 대상이 됨"
    )
    
    uploaded_file = st.file_uploader("직원 명단 (Excel) 업로드", type=["xlsx", "xls"])
    
    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        df.columns = [str(c).strip() for c in df.columns]
        
        required_cols = ['이름', '현재소속매장', '현재 매장 근무 시작일', '주방 숙련도', '홀 숙련도', '현재 직무', '가능 직무']
        for col in required_cols:
            if col not in df.columns:
                st.error(f"엑셀 파일에 '{col}' 열이 없습니다.")
                st.stop()
                
        if '이동가능여부' not in df.columns:
            df['이동가능여부'] = '가능'
            
        base_date = datetime.date.today()
        df['현재 근무개월수'] = df['현재 매장 근무 시작일'].apply(lambda x: calculate_months_between(x, base_date))
        
        df['현재 직무'] = df['현재 직무'].str.strip()
        df['가능 직무'] = df['가능 직무'].str.strip()
        
        invalid_jobs = df[~df['현재 직무'].isin(['주방', '홀'])]
        if len(invalid_jobs) > 0:
            st.warning(f"⚠️ 다음 직원의 '현재 직무'가 '주방' 또는 '홀'이 아닙니다: {', '.join(invalid_jobs['이름'].tolist())}")
            st.info("'현재 직무' 컬럼은 정확히 '주방' 또는 '홀'로만 입력해주세요.")
        
        df['가능직무목록'] = df['가능 직무'].apply(
            lambda x: [j.strip() for j in str(x).replace('/', ',').split(',') 
                      if j.strip() in ['주방', '홀']]
        )
        
        df['주방 숙련도'] = df['주방 숙련도'].apply(lambda x: safe_float(x))
        df['홀 숙련도'] = df['홀 숙련도'].apply(lambda x: safe_float(x))
        
        stores = [s for s in df['현재소속매장'].dropna().unique() if str(s).strip() != ""]
        
        kitchen_skills = df['주방 숙련도'].values
        hall_skills = df['홀 숙련도'].values
        
        kitchen_upper = np.percentile(kitchen_skills, 70)
        kitchen_lower = np.percentile(kitchen_skills, 30)
        
        hall_upper = np.percentile(hall_skills, 70)
        hall_lower = np.percentile(hall_skills, 30)
        
        st.sidebar.markdown("---")
        st.sidebar.caption(f"📊 숙련도 분포 분석")
        st.sidebar.caption(f"주방 상위: {kitchen_upper:.1f}점 이상")
        st.sidebar.caption(f"주방 중위: {kitchen_lower:.1f}~{kitchen_upper:.1f}점")
        st.sidebar.caption(f"주방 하위: {kitchen_lower:.1f}점 이하")
        st.sidebar.caption(f"홀 상위: {hall_upper:.1f}점 이상")
        st.sidebar.caption(f"홀 중위: {hall_lower:.1f}~{hall_upper:.1f}점")
        st.sidebar.caption(f"홀 하위: {hall_lower:.1f}점 이하")
        
        st.sidebar.subheader("🎯 매장별 목표 숙련도 등급")
        st.sidebar.caption("각 매장×직무에 필요한 숙련도 수준")
        
        target_levels = {}
        level_options = ["상위 (숙련도 높음)", "중위 (보통)", "하위 (낮음)"]
        level_map = {"상위 (숙련도 높음)": "상위", "중위 (보통)": "중위", "하위 (낮음)": "하위"}
        
        for store in stores:
            with st.sidebar.expander(f"📍 {store}"):
                kitchen_level = st.selectbox(
                    "주방 목표 등급",
                    options=level_options,
                    index=1,
                    key=f"level_kitchen_{store}"
                )
                hall_level = st.selectbox(
                    "홀 목표 등급",
                    options=level_options,
                    index=1,
                    key=f"level_hall_{store}"
                )
                target_levels[(store, '주방')] = level_map[kitchen_level]
                target_levels[(store, '홀')] = level_map[hall_level]
        
        st.sidebar.subheader("🏪 매장별 직무 정원")
        store_job_demands = {}
        for store in stores:
            with st.sidebar.expander(f"📍 {store} 인원"):
                current_kitchen = len(df[(df['현재소속매장'] == store) & (df['현재 직무'].str.contains('주방', na=False))])
                current_hall = len(df[(df['현재소속매장'] == store) & (df['현재 직무'].str.contains('홀', na=False))])
                
                kitchen_demand = st.number_input(
                    "주방 인원",
                    min_value=0, max_value=20, value=max(1, current_kitchen),
                    key=f"demand_kitchen_{store}"
                )
                hall_demand = st.number_input(
                    "홀 인원",
                    min_value=0, max_value=20, value=max(1, current_hall),
                    key=f"demand_hall_{store}"
                )
                store_job_demands[(store, '주방')] = kitchen_demand
                store_job_demands[(store, '홀')] = hall_demand
        
        st.subheader("📊 입력 데이터")
        st.dataframe(df[['이름', '현재소속매장', '현재 직무', '가능 직무', '현재 근무개월수', '주방 숙련도', '홀 숙련도', '이동가능여부']], use_container_width=True)
        
        if st.button("✨ 중장기 순환 계획 생성", type="primary"):
            
            rotation_points = list(range(first_rotation_month, planning_period_months + 1, 3))
            if planning_period_months not in rotation_points:
                rotation_points.append(planning_period_months)
            
            st.subheader(f"📅 순환 계획 타임라인 ({len(rotation_points)}회 순환)")
            timeline_info = " → ".join([f"{m}개월" for m in rotation_points])
            st.info(f"순환 시점: {timeline_info}")
            
            current_assignment = {}
            for _, row in df.iterrows():
                current_assignment[row['이름']] = (row['현재소속매장'], row['현재 직무'])
            
            rotation_history = []
            
            for rotation_idx, target_month in enumerate(rotation_points):
                st.markdown(f"### 🔄 {rotation_idx + 1}차 순환 ({target_month}개월 후)")
                
                rotation_results = []
                
                employees = df['이름'].tolist()
                jobs = ['주방', '홀']
                
                model = pulp.LpProblem(f"Rotation_{rotation_idx}", pulp.LpMinimize)
                x = pulp.LpVariable.dicts("assign", (employees, stores, jobs), cat="Binary")
                
                avg_skill = pulp.LpVariable.dicts("avg_skill", [(s, j) for s in stores for j in jobs], lowBound=0, cat="Continuous")
                
                def get_target_score(job, level):
                    if job == '주방':
                        if level == '상위': return kitchen_upper
                        elif level == '하위': return kitchen_lower
                        else: return (kitchen_upper + kitchen_lower) / 2
                    else:
                        if level == '상위': return hall_upper
                        elif level == '하위': return hall_lower
                        else: return (hall_upper + hall_lower) / 2
                
                skill_dev = pulp.LpVariable.dicts("skill_dev", [(s, j) for s in stores for j in jobs], lowBound=0, cat="Continuous")
                max_skill_dev = pulp.LpVariable("max_skill_dev", lowBound=0, cat="Continuous")
                
                for name in employees:
                    model += pulp.lpSum(x[name][s][j] for s in stores for j in jobs) == 1
                
                for s in stores:
                    for j in jobs:
                        model += pulp.lpSum(x[name][s][j] for name in employees) == store_job_demands.get((s, j), 0)
                
                for _, row in df.iterrows():
                    name = row['이름']
                    possible_jobs = row['가능직무목록']
                    for s in stores:
                        for j in jobs:
                            if j not in possible_jobs:
                                model += x[name][s][j] == 0
                
                for _, row in df.iterrows():
                    status = str(row['이동가능여부']).strip()
                    if '불' in status or status in ('N', 'n'):
                        name = row['이름']
                        if name in current_assignment:
                            current_store, current_job = current_assignment[name]
                            model += x[name][current_store][current_job] == 1
                
                for s in stores:
                    for j in jobs:
                        demand = store_job_demands.get((s, j), 0)
                        if demand == 0:
                            continue
                        
                        if j == '주방':
                            total_skill = pulp.lpSum(
                                x[row['이름']][s][j] * safe_float(row['주방 숙련도'])
                                for _, row in df.iterrows()
                            )
                        else:
                            total_skill = pulp.lpSum(
                                x[row['이름']][s][j] * safe_float(row['홀 숙련도'])
                                for _, row in df.iterrows()
                            )
                        
                        model += avg_skill[(s, j)] * demand == total_skill
                
                for s in stores:
                    for j in jobs:
                        if store_job_demands.get((s, j), 0) == 0:
                            continue
                        level = target_levels.get((s, j), '중위')
                        target_score = get_target_score(j, level)
                        
                        model += skill_dev[(s, j)] >= avg_skill[(s, j)] - target_score
                        model += skill_dev[(s, j)] >= target_score - avg_skill[(s, j)]
                        model += max_skill_dev >= skill_dev[(s, j)]
                
                objective = []
                
                objective.append(10000 * max_skill_dev)
                objective.append(5000 * pulp.lpSum(skill_dev[(s, j)] for s in stores for j in jobs 
                                                    if store_job_demands.get((s, j), 0) > 0))
                
                for name in employees:
                    if name in current_assignment:
                        current_store, current_job = current_assignment[name]
                    else:
                        emp_row = df[df['이름'] == name].iloc[0]
                        current_store = emp_row['현재소속매장']
                        current_job = emp_row['현재 직무']
                    
                    months_at_current = df[df['이름'] == name]['현재 근무개월수'].values[0] + target_month
                    
                    for s in stores:
                        for j in jobs:
                            if months_at_current >= 6:
                                if (s, j) == (current_store, current_job):
                                    cost = months_at_current * 10
                                else:
                                    cost = -months_at_current * 5
                            else:
                                if (s, j) != (current_store, current_job):
                                    cost = 100
                                else:
                                    cost = 0
                            objective.append(x[name][s][j] * cost)
                
                model += pulp.lpSum(objective)
                model.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=60))
                
                new_assignment = {}
                rotation_results = []
                for name in employees:
                    emp_row = df[df['이름'] == name].iloc[0]
                    for s in stores:
                        for j in jobs:
                            if pulp.value(x[name][s][j]) == 1:
                                new_assignment[name] = (s, j)
                                
                                if name in current_assignment:
                                    prev_store, prev_job = current_assignment[name]
                                else:
                                    prev_store = emp_row['현재소속매장']
                                    prev_job = emp_row['현재 직무']
                                
                                rotation_results.append({
                                    "이름": name,
                                    "기존 매장": prev_store,
                                    "기존 직무": prev_job,
                                    "배치 매장": s,
                                    "배치 직무": j,
                                    "이동": "🔄" if (s, j) != (prev_store, prev_job) else "유지",
                                    "주방": f"{emp_row['주방 숙련도']:.1f}",
                                    "홀": f"{emp_row['홀 숙련도']:.1f}"
                                })
                
                res_df = pd.DataFrame(rotation_results)
                
                store_job_stats = []
                for s in stores:
                    for j in jobs:
                        store_job_emps = res_df[(res_df['배치 매장'] == s) & (res_df['배치 직무'] == j)]
                        if len(store_job_emps) > 0:
                            if j == '주방':
                                avg_actual = store_job_emps['주방'].astype(float).mean()
                            else:
                                avg_actual = store_job_emps['홀'].astype(float).mean()
                            
                            level = target_levels.get((s, j), '중위')
                            target_score = get_target_score(j, level)
                            
                            store_job_stats.append({
                                "매장": s,
                                "직무": j,
                                "인원": len(store_job_emps),
                                "목표": level,
                                "실제 평균": f"{avg_actual:.1f}",
                                "목표 기준": f"{target_score:.1f}",
                                "편차": f"{abs(avg_actual - target_score):.1f}"
                            })
                
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.caption("발령 명단")
                    st.dataframe(res_df.sort_values(by=["배치 매장", "배치 직무"]), use_container_width=True, hide_index=True)
                
                with col2:
                    st.caption("매장×직무별 숙련도 현황")
                    st.dataframe(pd.DataFrame(store_job_stats), use_container_width=True, hide_index=True)
                
                current_assignment = new_assignment
                
                rotation_history.append({
                    'month': target_month,
                    'results': res_df.copy(),
                    'stats': pd.DataFrame(store_job_stats)
                })
            
            st.success(f"✅ {planning_period_months}개월 중장기 순환 계획 완료 ({len(rotation_points)}회 순환)")
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                for idx, history in enumerate(rotation_history):
                    sheet_name = f"{history['month']}개월_발령"
                    history['results'].to_excel(writer, sheet_name=sheet_name, index=False)
                    history['stats'].to_excel(writer, sheet_name=f"{history['month']}개월_통계", index=False)
            
            st.download_button(
                "📥 전체 순환 계획 엑셀 다운로드",
                output.getvalue(),
                f"중장기순환계획_{planning_period_months}개월.xlsx"
            )


else:  # 피로도 분석 (최근 스케줄 분석)":
    st.title("🥐 성심당 외식사업부 인력 운영 최적화 시스템")
    st.header("😴 직원 피로도 분석")
    st.info("최근 3~4개월 스케줄을 분석하여 직원별 피로도를 계산합니다.")
    
    import numpy as np
    from collections import defaultdict
    
    st.subheader("📂 최근 스케줄 파일 업로드")
    st.caption("여러 파일을 선택할 수 있습니다 (Ctrl+클릭). 운영 모드에서 생성된 엑셀 파일을 업로드하세요.")
    
    uploaded_files = st.file_uploader(
        "스케줄 파일 업로드 (여러 개 가능)", 
        type=["xlsx", "xls"],
        accept_multiple_files=True
    )
    
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)}개 파일 업로드 완료")
        
        # 전체 데이터 통합
        all_schedules = []
        file_info = []
        
        for uploaded_file in uploaded_files:
            try:
                # 먼저 일반적인 형식으로 읽기 시도
                df = pd.read_excel(uploaded_file, sheet_name=0)
                
                # 형식 판단: 컬럼에 숫자가 있으면 표준 형식
                has_date_cols = any(str(c).isdigit() for c in df.columns)
                
                if not has_date_cols:
                    # 변환 필요: 원본 형식
                    df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
                    
                    # 헤더 행 찾기 (직급, 이름이 있는 행)
                    header_row = None
                    for idx in range(min(10, len(df_raw))):
                        row = df_raw.iloc[idx]
                        if any('직급' in str(val) for val in row if pd.notna(val)) and \
                           any('이름' in str(val) for val in row if pd.notna(val)):
                            header_row = idx
                            break
                    
                    if header_row is None:
                        st.error(f"❌ {uploaded_file.name}: 헤더를 찾을 수 없습니다.")
                        continue
                    
                    # 데이터 시작 행
                    data_start_row = header_row + 2
                    
                    # 헤더 추출
                    headers = df_raw.iloc[header_row].tolist()
                    
                    # 날짜 컬럼 인덱스 찾기
                    date_indices = []
                    for i, h in enumerate(headers):
                        try:
                            if pd.notna(h) and str(h).strip().isdigit():
                                date_indices.append(i)
                        except:
                            pass
                    
                    # 이름 컬럼 인덱스 찾기
                    name_col_idx = None
                    for i, h in enumerate(headers):
                        if pd.notna(h) and '이름' in str(h):
                            name_col_idx = i
                            break
                    
                    if name_col_idx is None or len(date_indices) == 0:
                        st.error(f"❌ {uploaded_file.name}: 이름 또는 날짜 컬럼을 찾을 수 없습니다.")
                        continue
                    
                    # 데이터 변환
                    data_rows = []
                    for idx in range(data_start_row, len(df_raw)):
                        row = df_raw.iloc[idx]
                        name = row[name_col_idx]
                        
                        if pd.isna(name) or str(name).strip() == '':
                            continue
                        
                        row_data = {'이름': str(name).strip()}
                        
                        for date_idx in date_indices:
                            date_num = str(headers[date_idx]).strip()
                            work_val = row[date_idx]
                            row_data[date_num] = str(work_val).strip() if pd.notna(work_val) else ''
                        
                        data_rows.append(row_data)
                    
                    df = pd.DataFrame(data_rows)
                    st.success(f"✅ {uploaded_file.name}: 자동 변환 완료")
                
                # 파일 정보 수집
                total_days = len([c for c in df.columns if str(c).isdigit()])
                file_info.append({
                    "파일명": uploaded_file.name,
                    "근무일수": total_days,
                    "직원수": len(df)
                })
                
                all_schedules.append({
                    'filename': uploaded_file.name,
                    'data': df
                })
                
            except Exception as e:
                st.error(f"❌ {uploaded_file.name} 읽기 실패: {str(e)}")
        
        # 업로드된 파일 목록
        st.dataframe(pd.DataFrame(file_info), use_container_width=True, hide_index=True)
        
        if st.button("📊 피로도 분석 시작", type="primary"):
            
            # 직원별 데이터 수집
            employee_data = defaultdict(lambda: {
                'total_days': 0,
                'weekend_days': 0,
                'time_counts': defaultdict(int),
                'consecutive_work': [],
                'current_streak': 0,
                'holidays': 0
            })
            
            # 각 파일 분석
            for schedule in all_schedules:
                df = schedule['data']
                
                # 이름 컬럼 찾기
                name_col = None
                for col in df.columns:
                    if '이름' in str(col) or 'name' in str(col).lower():
                        name_col = col
                        break
                
                if name_col is None:
                    st.warning(f"⚠️ {schedule['filename']}: 이름 컬럼을 찾을 수 없습니다.")
                    continue
                
                # 날짜 컬럼 추출 (숫자로 된 컬럼)
                date_cols = [c for c in df.columns if str(c).isdigit()]
                
                for _, row in df.iterrows():
                    name = str(row[name_col]).strip()
                    if name == 'nan' or name == '':
                        continue
                    
                    for date_col in date_cols:
                        day_val = str(row[date_col]).strip().upper()
                        
                        if day_val in ['휴', '연', 'NAN', '']:
                            # 휴무/연차
                            if employee_data[name]['current_streak'] > 0:
                                employee_data[name]['consecutive_work'].append(
                                    employee_data[name]['current_streak']
                                )
                                employee_data[name]['current_streak'] = 0
                            
                            if day_val == '연':
                                employee_data[name]['holidays'] += 1
                        else:
                            # 근무
                            employee_data[name]['total_days'] += 1
                            employee_data[name]['current_streak'] += 1
                            
                            # 타임 카운트
                            if day_val in ['A', 'B', 'C', 'D', 'E']:
                                employee_data[name]['time_counts'][day_val] += 1
                            
                            # 주말 여부 판단 (날짜가 필요하지만 일단 간단히 처리)
                            # 실제로는 년/월 정보가 필요함
                            try:
                                day_num = int(date_col)
                                # 임시: 토(6), 일(0)로 가정 (실제로는 datetime 필요)
                                # 여기서는 대략적으로 계산
                            except:
                                pass
                
                # 마지막 연속 근무 추가
                for name in employee_data:
                    if employee_data[name]['current_streak'] > 0:
                        employee_data[name]['consecutive_work'].append(
                            employee_data[name]['current_streak']
                        )
                        employee_data[name]['current_streak'] = 0
            
            # 피로도 계산
            fatigue_results = []
            
            st.caption(f"🔍 수집된 직원 수: {len(employee_data)}")
            
            for name, data in employee_data.items():
                if data['total_days'] == 0:
                    continue
                
                # 타임 편중도 계산
                time_counts = data['time_counts']
                total_time_work = sum(time_counts.values())
                
                if total_time_work > 0:
                    time_concentration = max(time_counts.values()) / total_time_work
                    dominant_time = max(time_counts, key=time_counts.get) if time_counts else '-'
                else:
                    time_concentration = 0
                    dominant_time = '-'
                
                # 연속 근무 최대값
                max_consecutive = max(data['consecutive_work']) if data['consecutive_work'] else 0
                avg_consecutive = np.mean(data['consecutive_work']) if data['consecutive_work'] else 0
                
                # 피로도 점수 계산 (0~10)
                fatigue_score = 0
                
                # 1. 타임 편중도 (0~3점)
                fatigue_score += time_concentration * 3.0
                
                # 2. 연속 근무 (0~3점)
                if max_consecutive >= 6:
                    fatigue_score += 3.0
                elif max_consecutive >= 5:
                    fatigue_score += 2.0
                elif max_consecutive >= 4:
                    fatigue_score += 1.0
                
                # 3. 평균 연속 근무 (0~2점)
                if avg_consecutive >= 4.5:
                    fatigue_score += 2.0
                elif avg_consecutive >= 3.5:
                    fatigue_score += 1.0
                
                # 4. D/E 타임 비율 (야간 피로, 0~2점)
                night_ratio = (time_counts.get('D', 0) + time_counts.get('E', 0)) / total_time_work if total_time_work > 0 else 0
                fatigue_score += night_ratio * 2.0
                
                # 등급 판정
                if fatigue_score >= 7:
                    level = "🔴 높음"
                elif fatigue_score >= 4:
                    level = "🟡 보통"
                else:
                    level = "🟢 낮음"
                
                fatigue_results.append({
                    "이름": name,
                    "총 근무일": data['total_days'],
                    "최대 연속근무": f"{max_consecutive}일",
                    "평균 연속근무": f"{avg_consecutive:.1f}일",
                    "주력 타임": f"{dominant_time} ({time_concentration*100:.0f}%)",
                    "야간 비율": f"{night_ratio*100:.0f}%",
                    "피로도": f"{fatigue_score:.1f}",
                    "등급": level
                })
            
            # 결과 표시
            if len(fatigue_results) == 0:
                st.error("❌ 피로도 분석 결과가 없습니다. 파일 형식을 확인해주세요.")
                st.info("파일에 이름 컬럼과 날짜별 근무 데이터가 있는지 확인해주세요.")
            else:
                result_df = pd.DataFrame(fatigue_results).sort_values(by="피로도", ascending=False)
                
                st.subheader("📊 피로도 분석 결과")
                st.dataframe(result_df, use_container_width=True, hide_index=True)
                
                # 요약 통계
                col1, col2, col3 = st.columns(3)
                with col1:
                    high_fatigue = len([r for r in fatigue_results if float(r['피로도']) >= 7])
                    st.metric("🔴 고피로 직원", f"{high_fatigue}명")
                with col2:
                    medium_fatigue = len([r for r in fatigue_results if 4 <= float(r['피로도']) < 7])
                    st.metric("🟡 보통 피로", f"{medium_fatigue}명")
                with col3:
                    low_fatigue = len([r for r in fatigue_results if float(r['피로도']) < 4])
                    st.metric("🟢 저피로 직원", f"{low_fatigue}명")
                
                # 경고 메시지
                if high_fatigue > 0:
                    st.warning(f"⚠️ 피로도가 높은 직원이 {high_fatigue}명 있습니다. 순환 배치 시 고려가 필요합니다.")
                
                # 엑셀 다운로드
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    result_df.to_excel(writer, sheet_name="피로도분석", index=False)
                
                st.download_button(
                    "📥 피로도 분석 결과 다운로드",
                    output.getvalue(),
                    "피로도분석결과.xlsx"
                )